from django.test import SimpleTestCase

from apps.reportes.dto.report_dto import (
    AlumnoAsistenciaRowDTO,
    AlumnoCalificacionRowDTO,
    AsistenciasReportDTO,
    CalificacionesReportDTO,
    CategoriaConcentradoDTO,
    ActividadColumnaDTO,
    MateriaEncabezadoDTO,
)
from apps.reportes.services import excel_generator, pdf_generator
from openpyxl import load_workbook


def _materia() -> MateriaEncabezadoDTO:
    return MateriaEncabezadoDTO(
        materia_id=1,
        nrc='12345',
        nombre='Servicios Web',
        seccion='001',
        clave='COMP-456',
        docente_nombre='Dr. Pérez',
        docente_id=10,
        periodo_id=2,
        periodo_nombre='2026-1',
        horario='Lun-Mie',
    )


class GeneratorTests(SimpleTestCase):
    def test_calificaciones_xlsx_contiene_promedios(self):
        dto = CalificacionesReportDTO(
            materia=_materia(),
            categorias=(
                CategoriaConcentradoDTO(
                    nombre='Exámenes',
                    porcentaje=40.0,
                    actividades=(ActividadColumnaDTO(actividad_id=1, nombre='Parcial 1'),),
                ),
            ),
            alumnos=(
                AlumnoCalificacionRowDTO(
                    alumno_id=1,
                    matricula='20240001',
                    nombre='Ana García',
                    calificaciones=(),
                    promedio_real=7.65,
                    promedio_redondeado=8,
                ),
            ),
        )
        buf = excel_generator.build_calificaciones_xlsx(dto)
        wb = load_workbook(buf)
        ws = wb.active
        self.assertIn('Promedio Real', [c.value for c in ws[6]])
        self.assertEqual(ws['A2'].value, 'Materia: Servicios Web')

    def test_asistencias_pdf_genera_bytes(self):
        dto = AsistenciasReportDTO(
            materia=_materia(),
            total_sesiones=10,
            porcentaje_asistencia_grupal=80.0,
            alumnos=(
                AlumnoAsistenciaRowDTO(
                    alumno_id=1,
                    matricula='20240001',
                    nombre='Ana García',
                    presentes=9,
                    retardos=1,
                    ausentes=0,
                    porcentaje_asistencia=90.0,
                ),
            ),
        )
        buf = pdf_generator.build_asistencias_pdf(dto)
        self.assertTrue(buf.getvalue().startswith(b'%PDF'))
