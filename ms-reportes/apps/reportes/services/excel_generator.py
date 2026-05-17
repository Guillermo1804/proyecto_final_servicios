"""Generación de reportes Excel (.xlsx) desde DTOs."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from apps.reportes.dto.report_dto import AsistenciasReportDTO, CalificacionesReportDTO


def _auto_width(ws, max_col: int, min_width: int = 10, max_width: int = 40) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        length = min_width
        for cell in ws[letter]:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = length + 2


def _actividad_columnas(dto: CalificacionesReportDTO) -> list[tuple[int, str]]:
    columnas: list[tuple[int, str]] = []
    for cat in dto.categorias:
        for act in cat.actividades:
            columnas.append((act.actividad_id, f'{act.nombre} ({cat.nombre})'))
    return columnas


def build_calificaciones_xlsx(dto: CalificacionesReportDTO) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Calificaciones'
    m = dto.materia

    ws['A1'] = 'BUAP — Facultad de Ciencias de la Computación'
    ws['A2'] = f'Materia: {m.nombre}'
    ws['A3'] = f'NRC: {m.nrc} | Sección: {m.seccion} | Clave: {m.clave}'
    ws['A4'] = f'Periodo: {m.periodo_nombre} | Docente: {m.docente_nombre} | Horario: {m.horario}'
    for row in range(1, 5):
        ws[f'A{row}'].font = Font(bold=True)

    columnas_act = _actividad_columnas(dto)
    header = ['Matrícula', 'Nombre'] + [label for _, label in columnas_act] + [
        'Promedio Real',
        'Promedio Redondeado',
    ]
    header_row = 6
    ws.append([])
    ws.append(header)
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    for alumno in dto.alumnos:
        notas = {c.actividad_id: c.calificacion for c in alumno.calificaciones}
        fila = [
            alumno.matricula,
            alumno.nombre,
            *[notas.get(act_id, '') for act_id, _ in columnas_act],
            alumno.promedio_real,
            alumno.promedio_redondeado,
        ]
        ws.append(fila)

    ws.freeze_panes = ws[f'A{header_row + 1}']
    _auto_width(ws, len(header))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_asistencias_xlsx(dto: AsistenciasReportDTO) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Asistencias'
    m = dto.materia

    ws['A1'] = 'BUAP — Facultad de Ciencias de la Computación'
    ws['A2'] = f'Materia: {m.nombre}'
    ws['A3'] = f'NRC: {m.nrc} | Sección: {m.seccion} | Clave: {m.clave}'
    ws['A4'] = (
        f'Periodo: {m.periodo_nombre} | Docente: {m.docente_nombre} | '
        f'Total sesiones: {dto.total_sesiones} | Asistencia grupal: {dto.porcentaje_asistencia_grupal:.1f}%'
    )
    for row in range(1, 5):
        ws[f'A{row}'].font = Font(bold=True)

    header_row = 6
    ws.append([])
    header = [
        'Matrícula',
        'Nombre',
        'Presentes',
        'Retardos',
        'Ausentes',
        '% Asistencia',
    ]
    ws.append(header)
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    for alumno in dto.alumnos:
        ws.append(
            [
                alumno.matricula,
                alumno.nombre,
                alumno.presentes,
                alumno.retardos,
                alumno.ausentes,
                alumno.porcentaje_asistencia,
            ]
        )

    ws.freeze_panes = ws[f'A{header_row + 1}']
    _auto_width(ws, len(header))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
