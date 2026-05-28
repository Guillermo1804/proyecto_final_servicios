import logging
from io import BytesIO
from decimal import Decimal
import unicodedata

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import serializers, status
from rest_framework.decorators import api_view

from apps.core.models import EstadoMateria, Ponderacion, Actividad, Calificacion
from apps.core.serializers import (
    PonderacionBatchSerializer,
    PonderacionOutputSerializer,
    ActividadInputSerializer,
    ActividadOutputSerializer,
    CalificacionInputSerializer,
    CalificacionOutputSerializer,
)
from django.conf import settings

from apps.core.event_bus.publishers import (
    publish_actividad_created,
    publish_calificacion_updated,
    publish_concentrado_calculado,
    publish_materia_calificaciones_cerradas,
)
from apps.core.services import calcular_promedio_ponderado, obtener_concentrado_materia, redondear_institucional
from apps.core.projection_access import (
    alumno_puede_ver_materia,
    ensure_alumno_en_materia_projection,
    get_materia_local,
    is_alumno_en_materia_local,
    list_alumnos_materia_local,
    usuario_puede_gestionar_materia,
)
from utils.jwt_local import validate_access_token
from utils.responses import error_response, success_response

logger = logging.getLogger(__name__)


def _extract_bearer_token(request):
    authorization = request.headers.get('Authorization') or request.META.get('HTTP_AUTHORIZATION', '')
    if not authorization:
        return None
    if authorization.lower().startswith('bearer '):
        return authorization.split(None, 1)[1].strip()
    return authorization.strip()


def _authorize_materia_management(request, materia_id):
    token = _extract_bearer_token(request)
    if not token:
        return error_response('No se proporcionó token de autorización.', status=status.HTTP_401_UNAUTHORIZED)

    try:
        auth_response = validate_access_token(token)
    except ValueError:
        return error_response('Token inválido o expirado.', status=status.HTTP_401_UNAUTHORIZED)
    except Exception as exc:
        logger.exception('Error validando token para materia %s', materia_id)
        return error_response(f'No se pudo validar el token: {exc}', status=status.HTTP_503_SERVICE_UNAVAILABLE)

    role = (auth_response.rol or '').lower()
    if role not in ('admin', 'docente'):
        return error_response('No tienes permisos para gestionar esta materia.', status=status.HTTP_403_FORBIDDEN)

    try:
        materia = get_materia_local(materia_id)
    except LookupError:
        return error_response('La materia no existe en proyección local.', status=status.HTTP_404_NOT_FOUND)

    if not usuario_puede_gestionar_materia(
        usuario_id=auth_response.user_id,
        usuario_email=getattr(auth_response, 'email', '') or '',
        usuario_rol=auth_response.rol,
        materia=materia,
    ):
        return error_response('No tienes permisos para gestionar esta materia.', status=status.HTTP_403_FORBIDDEN)

    return None


def _authorize_materia_read(request, materia_id):
    """GET: admin, docente de la materia o alumno inscrito (proyección local)."""
    token = _extract_bearer_token(request)
    if not token:
        return error_response('No se proporcionó token de autorización.', status=status.HTTP_401_UNAUTHORIZED)

    try:
        auth_response = validate_access_token(token)
    except ValueError:
        return error_response('Token inválido o expirado.', status=status.HTTP_401_UNAUTHORIZED)
    except Exception as exc:
        logger.exception('Error validando token para lectura materia %s', materia_id)
        return error_response(f'No se pudo validar el token: {exc}', status=status.HTTP_503_SERVICE_UNAVAILABLE)

    role = (auth_response.rol or '').lower()
    if role == 'admin':
        return None

    try:
        get_materia_local(materia_id)
    except LookupError:
        return error_response('La materia no existe en proyección local.', status=status.HTTP_404_NOT_FOUND)

    if role == 'docente':
        try:
            materia = get_materia_local(materia_id)
        except LookupError:
            return error_response('La materia no existe en proyección local.', status=status.HTTP_404_NOT_FOUND)
        if usuario_puede_gestionar_materia(
            usuario_id=auth_response.user_id,
            usuario_email=getattr(auth_response, 'email', '') or '',
            usuario_rol=auth_response.rol,
            materia=materia,
        ):
            return None
        return error_response('No tienes permisos para ver esta materia.', status=status.HTTP_403_FORBIDDEN)

    if role == 'alumno':
        if alumno_puede_ver_materia(
            materia_id,
            usuario_email=getattr(auth_response, 'email', '') or '',
        ):
            return None
        return error_response('No estás inscrito en esta materia.', status=status.HTTP_403_FORBIDDEN)

    return error_response('No tienes permisos para ver esta materia.', status=status.HTTP_403_FORBIDDEN)


def _parse_ponderaciones_payload(request):
    payload = request.data
    if isinstance(payload, list):
        return {'ponderaciones': payload}
    if isinstance(payload, dict) and 'ponderaciones' in payload:
        return {'ponderaciones': payload['ponderaciones']}
    return payload


def _normalize_excel_header(value):
    if value is None:
        return ''

    normalized = unicodedata.normalize('NFKD', str(value)).encode('ascii', 'ignore').decode('ascii')
    return normalized.strip().casefold()


def _set_ponderaciones(materia_id, ponderaciones_data):
    serializer = PonderacionBatchSerializer(data={'ponderaciones': ponderaciones_data})
    serializer.is_valid(raise_exception=True)

    with transaction.atomic():
        Ponderacion.objects.filter(materia_id=materia_id).delete()
        Ponderacion.objects.bulk_create([
            Ponderacion(
                materia_id=materia_id,
                nombre_categoria=item['nombre_categoria'],
                porcentaje=item['porcentaje'],
            )
            for item in serializer.validated_data['ponderaciones']
        ])

    ponderaciones_guardadas = Ponderacion.objects.filter(materia_id=materia_id).order_by('id')
    total = sum((item.porcentaje for item in ponderaciones_guardadas), 0)
    return ponderaciones_guardadas, total


def _read_ponderaciones_from_excel(uploaded_file):
    try:
        workbook = load_workbook(filename=BytesIO(uploaded_file.read()), data_only=True)
    except Exception as exc:
        raise serializers.ValidationError({'archivo': [f'No se pudo leer el archivo Excel: {exc}']}) from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise serializers.ValidationError({'archivo': ['El archivo Excel está vacío.']})

    headers = [_normalize_excel_header(value) for value in rows[0]]
    header_map = {header: index for index, header in enumerate(headers) if header}

    nombre_key = next((key for key in ('nombre_categoria', 'categoria', 'nombre') if key in header_map), None)
    porcentaje_key = next((key for key in ('porcentaje', 'peso') if key in header_map), None)
    if nombre_key is None or porcentaje_key is None:
        raise serializers.ValidationError({
            'archivo': ['Las columnas requeridas son nombre_categoria/categoria/nombre y porcentaje/peso.'],
        })

    ponderaciones = []
    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        nombre_categoria = row[header_map[nombre_key]] if header_map[nombre_key] < len(row) else None
        porcentaje = row[header_map[porcentaje_key]] if header_map[porcentaje_key] < len(row) else None
        ponderaciones.append({
            'nombre_categoria': nombre_categoria,
            'porcentaje': porcentaje,
        })

    return ponderaciones


def _read_calificaciones_from_excel(uploaded_file):
    try:
        workbook = load_workbook(filename=BytesIO(uploaded_file.read()), data_only=True)
    except Exception as exc:
        raise serializers.ValidationError({'archivo': [f'No se pudo leer el archivo Excel: {exc}']}) from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise serializers.ValidationError({'archivo': ['El archivo Excel está vacío.']})

    headers = [_normalize_excel_header(value) for value in rows[0]]
    header_map = {header: index for index, header in enumerate(headers) if header}

    alumno_key = next((key for key in ('matricula', 'alumno_id', 'alumno') if key in header_map), None)
    actividad_key = next((key for key in ('actividad_id', 'actividad') if key in header_map), None)
    calif_key = next((key for key in ('calificacion', 'nota', 'calif') if key in header_map), None)

    if alumno_key is None or actividad_key is None or calif_key is None:
        raise serializers.ValidationError({
            'archivo': ['Las columnas requeridas son matricula/alumno_id, actividad_id y calificacion/nota.'],
        })

    filas = []
    for idx, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None for cell in row):
            continue
        alumno = row[header_map[alumno_key]] if header_map[alumno_key] < len(row) else None
        actividad = row[header_map[actividad_key]] if header_map[actividad_key] < len(row) else None
        calificacion = row[header_map[calif_key]] if header_map[calif_key] < len(row) else None
        filas.append({
            'fila': idx,
            'alumno_id': alumno,
            'actividad_id': actividad,
            'calificacion': calificacion,
        })

    return filas


def _build_concentrado_rest_payload(materia_id):
    concentrado_local = obtener_concentrado_materia(materia_id)
    if concentrado_local is None:
        return None

    alumnos_en_materia = list_alumnos_materia_local(materia_id)
    alumnos_index = {alumno.id: alumno for alumno in alumnos_en_materia}

    actividades_por_alumno = {}
    calificaciones_qs = Calificacion.objects.filter(actividad__ponderacion__materia_id=materia_id).select_related('actividad__ponderacion', 'actividad')
    for calificacion in calificaciones_qs.order_by('actividad__ponderacion__id', 'actividad__id'):
        actividades_por_alumno.setdefault(calificacion.alumno_id, []).append(calificacion)

    alumno_ids = [alumno.id for alumno in alumnos_en_materia]
    if not alumno_ids:
        alumno_ids = sorted(set(actividades_por_alumno.keys()))

    categorias = []
    for ponderacion, actividades in concentrado_local['categorias']:
        categorias.append({
            'nombre': ponderacion.nombre_categoria,
            'porcentaje': str(ponderacion.porcentaje),
            'actividades': [
                {'id': actividad.id, 'nombre': actividad.nombre}
                for actividad in actividades
            ],
        })

    alumnos = []
    for alumno_id in alumno_ids:
        alumno_ref = alumnos_index.get(alumno_id)
        calificaciones_alumno = actividades_por_alumno.get(alumno_id, [])
        promedio_real = calcular_promedio_ponderado(alumno_id, materia_id)
        promedio_real = promedio_real.quantize(Decimal('0.01'))
        alumnos.append({
            'alumno_id': alumno_id,
            'matricula': alumno_ref.matricula if alumno_ref else str(alumno_id),
            'nombre': alumno_ref.nombre if alumno_ref else '',
            'calificaciones': [
                {
                    'actividad_id': item.actividad_id,
                    'actividad_nombre': item.actividad.nombre,
                    'categoria': item.actividad.ponderacion.nombre_categoria,
                    'calificacion': str(item.calificacion),
                }
                for item in calificaciones_alumno
            ],
            'promedio_real': str(promedio_real),
            'promedio_redondeado': redondear_institucional(promedio_real),
        })

    return {
        'materia_id': materia_id,
        'categorias': categorias,
        'alumnos': alumnos,
    }


@api_view(['GET', 'POST', 'PUT'])
def ponderaciones(request, materia_id: int):
    if materia_id <= 0:
        return error_response('materia_id inválido', status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'GET':
        auth_error = _authorize_materia_read(request, materia_id)
        if auth_error is not None:
            return auth_error
        ponderaciones_qs = Ponderacion.objects.filter(materia_id=materia_id).order_by('id')
        serializer = PonderacionOutputSerializer(ponderaciones_qs, many=True)
        total = sum((item.porcentaje for item in ponderaciones_qs), 0)
        return success_response(
            {
                'materia_id': materia_id,
                'ponderaciones': serializer.data,
                'total': str(total),
            },
            message='Ponderaciones obtenidas correctamente',
            status=status.HTTP_200_OK,
        )

    auth_error = _authorize_materia_management(request, materia_id)
    if auth_error is not None:
        return auth_error

    payload = _parse_ponderaciones_payload(request)
    if not isinstance(payload, dict) or 'ponderaciones' not in payload:
        return error_response('Se esperaba un objeto con la clave "ponderaciones".', status=status.HTTP_400_BAD_REQUEST)

    try:
        ponderaciones_guardadas, total = _set_ponderaciones(materia_id, payload['ponderaciones'])
    except serializers.ValidationError as exc:
        return error_response('Validación de ponderaciones fallida.', errors=exc.detail, status=status.HTTP_400_BAD_REQUEST)

    serializer = PonderacionOutputSerializer(ponderaciones_guardadas, many=True)
    return success_response(
        {
            'materia_id': materia_id,
            'ponderaciones': serializer.data,
            'total': str(total),
        },
        message='Ponderaciones guardadas correctamente',
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
def importar_ponderaciones(request, materia_id: int):
    if materia_id <= 0:
        return error_response('materia_id inválido', status=status.HTTP_400_BAD_REQUEST)

    auth_error = _authorize_materia_management(request, materia_id)
    if auth_error is not None:
        return auth_error

    uploaded_file = request.FILES.get('archivo') or request.FILES.get('file')
    if uploaded_file is None:
        return error_response('Debes enviar un archivo Excel en el campo "archivo".', status=status.HTTP_400_BAD_REQUEST)

    try:
        ponderaciones_excel = _read_ponderaciones_from_excel(uploaded_file)
        ponderaciones_guardadas, total = _set_ponderaciones(materia_id, ponderaciones_excel)
    except serializers.ValidationError as exc:
        return error_response('Validación de importación fallida.', errors=exc.detail, status=status.HTTP_400_BAD_REQUEST)

    serializer = PonderacionOutputSerializer(ponderaciones_guardadas, many=True)
    return success_response(
        {
            'materia_id': materia_id,
            'ponderaciones': serializer.data,
            'total': str(total),
            'importadas': len(serializer.data),
        },
        message='Ponderaciones importadas correctamente',
        status=status.HTTP_200_OK,
    )


@api_view(['GET', 'POST'])
def actividades(request):
    """
    GET /actividades?materia=:id → Listar actividades agrupadas por categoría
    POST /actividades → Crear nueva actividad
    """
    if request.method == 'POST':
        serializer = ActividadInputSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ponderacion_id = serializer.validated_data['ponderacion_id']
        try:
            ponderacion = Ponderacion.objects.get(id=ponderacion_id)
        except Ponderacion.DoesNotExist:
            return error_response('La ponderación no existe.', status=status.HTTP_404_NOT_FOUND)

        materia_id = ponderacion.materia_id
        auth_error = _authorize_materia_management(request, materia_id)
        if auth_error is not None:
            return auth_error

        try:
            actividad = Actividad.objects.create(
                ponderacion=ponderacion,
                nombre=serializer.validated_data['nombre'],
                descripcion=serializer.validated_data.get('descripcion', ''),
                fecha=serializer.validated_data.get('fecha'),
            )
            fecha_val = actividad.fecha.isoformat() if actividad.fecha else None
            publish_actividad_created(
                actividad_id=actividad.id,
                materia_id=materia_id,
                ponderacion_id=ponderacion.id,
                nombre=actividad.nombre,
                descripcion=actividad.descripcion,
                fecha=fecha_val,
                categoria=ponderacion.nombre_categoria,
            )
            output_serializer = ActividadOutputSerializer(actividad)
            return success_response(
                output_serializer.data,
                message='Actividad creada correctamente',
                status=status.HTTP_201_CREATED,
            )
        except Exception as exc:
            logger.exception('Error creando actividad')
            return error_response(f'Error al crear actividad: {exc}', status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    elif request.method == 'GET':
        materia_id = request.query_params.get('materia')
        if not materia_id:
            return error_response('El parámetro "materia" es obligatorio.', status=status.HTTP_400_BAD_REQUEST)

        try:
            materia_id = int(materia_id)
        except ValueError:
            return error_response('El parámetro "materia" debe ser un entero.', status=status.HTTP_400_BAD_REQUEST)

        auth_error = _authorize_materia_read(request, materia_id)
        if auth_error is not None:
            return auth_error

        ponderaciones = Ponderacion.objects.filter(materia_id=materia_id).order_by('id')
        resultado = []
        for pond in ponderaciones:
            actividades_qs = pond.actividades.all().order_by('id')
            output_serializer = ActividadOutputSerializer(actividades_qs, many=True)
            resultado.append({
                'categoria_nombre': pond.nombre_categoria,
                'categoria_porcentaje': str(pond.porcentaje),
                'actividades': output_serializer.data,
            })

        return success_response(
            {
                'materia_id': materia_id,
                'categorias': resultado,
            },
            message='Actividades obtenidas correctamente',
            status=status.HTTP_200_OK,
        )


@api_view(['PUT', 'DELETE'])
def editar_eliminar_actividad(request, actividad_id: int):
    """
    PUT /actividades/:id → actualizar actividad
    DELETE /actividades/:id → eliminar actividad (solo si no tiene calificaciones)
    """
    if actividad_id <= 0:
        return error_response('actividad_id inválido', status=status.HTTP_400_BAD_REQUEST)

    try:
        actividad = Actividad.objects.select_related('ponderacion').get(id=actividad_id)
    except Actividad.DoesNotExist:
        return error_response('La actividad no existe.', status=status.HTTP_404_NOT_FOUND)

    materia_id = actividad.ponderacion.materia_id
    auth_error = _authorize_materia_management(request, materia_id)
    if auth_error is not None:
        return auth_error

    if request.method == 'PUT':
        serializer = ActividadInputSerializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        if 'nombre' in serializer.validated_data:
            actividad.nombre = serializer.validated_data['nombre']
        if 'descripcion' in serializer.validated_data:
            actividad.descripcion = serializer.validated_data['descripcion']
        if 'fecha' in serializer.validated_data:
            actividad.fecha = serializer.validated_data['fecha']
        if 'ponderacion_id' in serializer.validated_data:
            try:
                nueva_pond = Ponderacion.objects.get(id=serializer.validated_data['ponderacion_id'])
                if nueva_pond.materia_id != materia_id:
                    return error_response(
                        'No puedes mover actividades entre materias.',
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                actividad.ponderacion = nueva_pond
            except Ponderacion.DoesNotExist:
                return error_response('La ponderación no existe.', status=status.HTTP_404_NOT_FOUND)

        actividad.save()
        output_serializer = ActividadOutputSerializer(actividad)
        return success_response(
            output_serializer.data,
            message='Actividad actualizada correctamente',
            status=status.HTTP_200_OK,
        )

    elif request.method == 'DELETE':
        calificaciones_count = actividad.calificaciones.count()
        if calificaciones_count > 0:
            return error_response(
                f'No se puede eliminar: la actividad tiene {calificaciones_count} calificación(es).',
                status=status.HTTP_409_CONFLICT,
            )

        actividad_id_deleted = actividad.id
        actividad.delete()
        return success_response(
            {'actividad_id': actividad_id_deleted},
            message='Actividad eliminada correctamente',
            status=status.HTTP_200_OK,
        )


@api_view(['POST'])
def crear_calificacion(request):
    """
    POST /calificaciones
    Crear o actualizar una calificación (upsert).
    Body: { actividad_id, alumno_id, calificacion }
    """
    serializer = CalificacionInputSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    actividad_id = serializer.validated_data['actividad_id']
    alumno_id = serializer.validated_data['alumno_id']
    calificacion = serializer.validated_data['calificacion']

    try:
        actividad = Actividad.objects.select_related('ponderacion').get(id=actividad_id)
    except Actividad.DoesNotExist:
        return error_response('La actividad no existe.', status=status.HTTP_404_NOT_FOUND)

    materia_id = actividad.ponderacion.materia_id

    auth_error = _authorize_materia_management(request, materia_id)
    if auth_error is not None:
        return auth_error

    # Verificar si la materia tiene lista impresa (bloquea nuevas calificaciones)
    try:
        estado = EstadoMateria.objects.get(materia_id=materia_id)
        if estado.lista_impresa:
            return error_response(
                'No se pueden crear/editar calificaciones: la lista ya fue impresa.',
                status=status.HTTP_409_CONFLICT,
            )
    except EstadoMateria.DoesNotExist:
        # Primera vez, no hay restricción
        pass

    matricula = serializer.validated_data.get('matricula', '') or ''
    nombre = serializer.validated_data.get('nombre', '') or ''
    email = serializer.validated_data.get('email', '') or ''
    if not ensure_alumno_en_materia_projection(
        materia_id,
        alumno_id,
        matricula=matricula,
        nombre=nombre,
        email=email,
    ):
        return error_response(
            'El alumno no está inscrito activo en la materia (proyección local).',
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        calif_obj, created = Calificacion.objects.update_or_create(
            actividad=actividad,
            alumno_id=alumno_id,
            defaults={'calificacion': calificacion},
        )
        publish_calificacion_updated(
            calificacion_id=calif_obj.id,
            actividad_id=actividad_id,
            alumno_id=alumno_id,
            materia_id=materia_id,
            calificacion=calificacion,
            created=created,
        )
        output_serializer = CalificacionOutputSerializer(calif_obj)
        action = 'creada' if created else 'actualizada'
        return success_response(
            output_serializer.data,
            message=f'Calificación {action} correctamente',
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )
    except Exception as exc:
        logger.exception('Error guardando calificación')
        return error_response(f'Error al guardar calificación: {exc}', status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def importar_calificaciones(request, materia_id: int):
    """
    POST /calificaciones/importar?materia=:id
    Importa un Excel con columnas: matricula/alumno_id, actividad_id, calificacion
    """
    if materia_id <= 0:
        return error_response('materia_id inválido', status=status.HTTP_400_BAD_REQUEST)

    auth_error = _authorize_materia_management(request, materia_id)
    if auth_error is not None:
        return auth_error

    try:
        estado = EstadoMateria.objects.get(materia_id=materia_id)
        if estado.lista_impresa:
            return error_response(
                'No se pueden importar calificaciones: la lista ya fue impresa.',
                status=status.HTTP_409_CONFLICT,
            )
    except EstadoMateria.DoesNotExist:
        pass

    uploaded_file = request.FILES.get('archivo') or request.FILES.get('file')
    if uploaded_file is None:
        return error_response('Debes enviar un archivo Excel en el campo "archivo".', status=status.HTTP_400_BAD_REQUEST)

    try:
        filas = _read_calificaciones_from_excel(uploaded_file)
    except serializers.ValidationError as exc:
        return error_response('Validación de importación fallida.', errors=exc.detail, status=status.HTTP_400_BAD_REQUEST)

    resumen = {
        'procesadas': 0,
        'importadas': 0,
        'fallos': 0,
        'errores': [],
    }

    for item in filas:
        resumen['procesadas'] += 1
        fila = item['fila']
        try:
            actividad_id = int(item['actividad_id']) if item['actividad_id'] is not None else None
        except Exception:
            resumen['fallos'] += 1
            resumen['errores'].append({'fila': fila, 'motivo': 'actividad_id inválido'})
            continue

        try:
            actividad = Actividad.objects.select_related('ponderacion').get(id=actividad_id)
        except Actividad.DoesNotExist:
            resumen['fallos'] += 1
            resumen['errores'].append({'fila': fila, 'motivo': 'actividad no existe'})
            continue

        if actividad.ponderacion.materia_id != materia_id:
            resumen['fallos'] += 1
            resumen['errores'].append({'fila': fila, 'motivo': 'actividad no pertenece a la materia'})
            continue

        try:
            alumno_id = int(item['alumno_id']) if item['alumno_id'] is not None else None
        except Exception:
            resumen['fallos'] += 1
            resumen['errores'].append({'fila': fila, 'motivo': 'alumno_id inválido'})
            continue

        if not is_alumno_en_materia_local(alumno_id, materia_id):
            resumen['fallos'] += 1
            resumen['errores'].append({'fila': fila, 'motivo': 'alumno no inscrito (proyección)'})
            continue

        try:
            cal_val = item['calificacion']
            cal_decimal = Decimal(str(cal_val))
            if cal_decimal < Decimal('0.00') or cal_decimal > Decimal('10.00'):
                raise ValueError('fuera de rango')
        except Exception:
            resumen['fallos'] += 1
            resumen['errores'].append({'fila': fila, 'motivo': 'calificacion inválida'})
            continue

        try:
            Calificacion.objects.update_or_create(
                actividad=actividad,
                alumno_id=alumno_id,
                defaults={'calificacion': cal_decimal},
            )
            resumen['importadas'] += 1
        except Exception as exc:
            resumen['fallos'] += 1
            resumen['errores'].append({'fila': fila, 'motivo': f'error guardando: {exc}'})

    resumen['ok'] = resumen['procesadas'] - resumen['fallos']

    return success_response(
        resumen,
        message='Importación procesada',
        status=status.HTTP_200_OK,
    )


@api_view(['GET'])
def concentrado(request, materia_id: int):
    if materia_id <= 0:
        return error_response('materia_id inválido', status=status.HTTP_400_BAD_REQUEST)

    auth_error = _authorize_materia_read(request, materia_id)
    if auth_error is not None:
        return auth_error

    try:
        payload = _build_concentrado_rest_payload(materia_id)
    except RuntimeError as exc:
        logger.exception('Error construyendo concentrado REST para materia %s', materia_id)
        return error_response(str(exc), status=status.HTTP_503_SERVICE_UNAVAILABLE)

    if payload is None:
        return error_response('La materia no tiene datos locales.', status=status.HTTP_404_NOT_FOUND)

    if getattr(settings, 'USE_EVENT_BUS', False) and payload.get('alumnos'):
        promedios = [Decimal(a['promedio_real']) for a in payload['alumnos']]
        promedio_grupal = (
            sum(promedios, Decimal('0')) / Decimal(len(promedios)) if promedios else Decimal('0')
        )
        try:
            materia_local = get_materia_local(materia_id)
            nrc, nombre = materia_local.nrc, materia_local.nombre
        except LookupError:
            nrc, nombre = '', f'Materia {materia_id}'
        publish_concentrado_calculado(
            materia_id=materia_id,
            total_alumnos=len(payload['alumnos']),
            promedio_grupal=float(promedio_grupal),
            nrc=nrc,
            materia_nombre=nombre,
        )

    return success_response(
        payload,
        message='Concentrado obtenido correctamente',
        status=status.HTTP_200_OK,
    )


@api_view(['PUT', 'DELETE'])
def editar_eliminar_calificacion(request, calificacion_id: int):
    """
    PUT /calificaciones/:id → actualizar calificación (bloqueado si lista impresa)
    DELETE /calificaciones/:id → eliminar calificación
    """
    if calificacion_id <= 0:
        return error_response('calificacion_id inválido', status=status.HTTP_400_BAD_REQUEST)

    try:
        calif = Calificacion.objects.select_related('actividad__ponderacion').get(id=calificacion_id)
    except Calificacion.DoesNotExist:
        return error_response('La calificación no existe.', status=status.HTTP_404_NOT_FOUND)

    materia_id = calif.actividad.ponderacion.materia_id

    auth_error = _authorize_materia_management(request, materia_id)
    if auth_error is not None:
        return auth_error

    if request.method == 'PUT':
        # Verificar que lista no esté impresa
        try:
            estado = EstadoMateria.objects.get(materia_id=materia_id)
            if estado.lista_impresa:
                return error_response(
                    'No se pueden editar calificaciones: la lista ya fue impresa.',
                    status=status.HTTP_409_CONFLICT,
                )
        except EstadoMateria.DoesNotExist:
            pass

        serializer = CalificacionInputSerializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        if 'calificacion' in serializer.validated_data:
            calif.calificacion = serializer.validated_data['calificacion']
            calif.save(update_fields=['calificacion'])

        output_serializer = CalificacionOutputSerializer(calif)
        return success_response(
            output_serializer.data,
            message='Calificación actualizada correctamente',
            status=status.HTTP_200_OK,
        )

    elif request.method == 'DELETE':
        calif_id = calif.id
        calif.delete()
        return success_response(
            {'calificacion_id': calif_id},
            message='Calificación eliminada correctamente',
            status=status.HTTP_200_OK,
        )


@api_view(['POST'])
def cerrar_materia(request, materia_id: int):
    """
    POST /materias/<id>/cerrar
    Marca la materia como cerrada y publica materia.calificaciones_cerradas.v1 (MS-6 vía bus).
    """
    if materia_id <= 0:
        return error_response('materia_id inválido', status=status.HTTP_400_BAD_REQUEST)

    auth_error = _authorize_materia_management(request, materia_id)
    if auth_error is not None:
        return auth_error

    try:
        materia_local = get_materia_local(materia_id)
    except LookupError:
        return error_response('La materia no existe en proyección local.', status=status.HTTP_404_NOT_FOUND)

    try:
        with transaction.atomic():
            estado, _ = EstadoMateria.objects.select_for_update().get_or_create(
                materia_id=materia_id,
                defaults={'cerrada': False},
            )
            if estado.cerrada:
                return error_response('La materia ya está cerrada', status=status.HTTP_400_BAD_REQUEST)

            estado.cerrada = True
            estado.fecha_cierre = timezone.now()
            estado.save(update_fields=['cerrada', 'fecha_cierre'])

            alumnos_payload = []
            for alumno in list_alumnos_materia_local(materia_id):
                promedio_real = calcular_promedio_ponderado(alumno.id, materia_id)
                alumnos_payload.append({
                    'alumno_id': alumno.id,
                    'email': alumno.email,
                    'matricula': alumno.matricula,
                    'nombre': alumno.nombre,
                    'promedio_real': float(promedio_real.quantize(Decimal('0.01'))),
                    'promedio_redondeado': redondear_institucional(promedio_real),
                })

            event_payload = {
                'materia_id': materia_id,
                'periodo_id': materia_local.periodo_id,
                'nrc': materia_local.nrc,
                'nombre': materia_local.nombre,
                'seccion': materia_local.seccion,
                'periodo_nombre': materia_local.periodo_nombre,
                'fecha_cierre': estado.fecha_cierre.isoformat(),
                'alumnos': alumnos_payload,
            }
            publish_materia_calificaciones_cerradas(event_payload)

        if getattr(settings, 'USE_EVENT_BUS', False):
            EstadoMateria.objects.filter(materia_id=materia_id).update(notificacion_enviada=True)

        return success_response(
            {
                'materia_id': materia_id,
                'cerrada': True,
                'evento_publicado': getattr(settings, 'USE_EVENT_BUS', False),
            },
            message='Materia cerrada correctamente',
            status=status.HTTP_200_OK,
        )
    except Exception as exc:
        logger.exception('Error al cerrar materia %s', materia_id)
        return error_response(f'Error al cerrar materia: {exc}', status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def imprimir_lista(request, materia_id: int):
    """
    POST /materias/<id>/imprimir-lista
    Marca `lista_impresa=True` para la materia (bloquea ediciones posteriores).
    """
    if materia_id <= 0:
        return error_response('materia_id inválido', status=status.HTTP_400_BAD_REQUEST)

    auth_error = _authorize_materia_management(request, materia_id)
    if auth_error is not None:
        return auth_error

    try:
        estado, _ = EstadoMateria.objects.get_or_create(
            materia_id=materia_id,
            defaults={'cerrada': False},
        )
        if estado.lista_impresa:
            return error_response('La lista ya fue impresa', status=status.HTTP_400_BAD_REQUEST)

        estado.lista_impresa = True
        estado.save(update_fields=['lista_impresa'])

        return success_response(
            {
                'materia_id': materia_id,
                'lista_impresa': True,
            },
            message='Lista marcada como impresa',
            status=status.HTTP_200_OK,
        )
    except Exception as exc:
        logger.exception('Error marcando lista impresa para materia %s', materia_id)
        return error_response(f'Error al marcar lista impresa: {exc}', status=status.HTTP_500_INTERNAL_SERVER_ERROR)
