from unittest.mock import patch
from types import SimpleNamespace
from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook

from django.test import TestCase, override_settings

from apps.core.models import (
    AlumnoMateriaProjection,
    EstadoMateria,
    MateriaProjection,
    Ponderacion,
    Actividad,
    Calificacion,
)
from apps.core.services import calcular_promedio_ponderado, obtener_estadisticas_materia, redondear_institucional


class ActividadTests(TestCase):
    def setUp(self):
        self.auth_headers = {'HTTP_AUTHORIZATION': 'Bearer token-valido'}
        self.ponderacion = None

    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_crear_actividad_ok(self, mock_validate, mock_materia):
        ponderacion = Ponderacion.objects.create(
            materia_id=10,
            nombre_categoria='Exámenes',
            porcentaje='40.00',
        )

        response = self.client.post(
            '/actividades',
            data={
                'ponderacion_id': ponderacion.id,
                'nombre': 'Examen Parcial 1',
                'descripcion': 'Primer parcial del curso',
                'fecha': '2024-10-15',
            },
            content_type='application/json',
            **self.auth_headers,
        )

        self.assertEqual(response.status_code, 201)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['data']['nombre'], 'Examen Parcial 1')
        mock_validate.assert_called_once_with('token-valido')

    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_listar_actividades_agrupadas(self, mock_validate, mock_materia):
        pond1 = Ponderacion.objects.create(
            materia_id=10,
            nombre_categoria='Exámenes',
            porcentaje='40.00',
        )
        pond2 = Ponderacion.objects.create(
            materia_id=10,
            nombre_categoria='Tareas',
            porcentaje='60.00',
        )
        Actividad.objects.create(ponderacion=pond1, nombre='Examen Parcial 1')
        Actividad.objects.create(ponderacion=pond1, nombre='Examen Parcial 2')
        Actividad.objects.create(ponderacion=pond2, nombre='Tarea 1')

        response = self.client.get('/actividades?materia=10', **self.auth_headers)
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(len(body['data']['categorias']), 2)
        self.assertEqual(len(body['data']['categorias'][0]['actividades']), 2)
        self.assertEqual(len(body['data']['categorias'][1]['actividades']), 1)

    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_eliminar_actividad_sin_calificaciones(self, mock_validate, mock_materia):
        pond = Ponderacion.objects.create(
            materia_id=10,
            nombre_categoria='Exámenes',
            porcentaje='100.00',
        )
        actividad = Actividad.objects.create(ponderacion=pond, nombre='Examen 1')

        response = self.client.delete(f'/actividades/{actividad.id}', **self.auth_headers)
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertFalse(Actividad.objects.filter(id=actividad.id).exists())

    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_eliminar_actividad_con_calificaciones_rechazada(self, mock_validate, mock_materia):
        from apps.core.models import Calificacion

        pond = Ponderacion.objects.create(
            materia_id=10,
            nombre_categoria='Exámenes',
            porcentaje='100.00',
        )
        actividad = Actividad.objects.create(ponderacion=pond, nombre='Examen 1')
        Calificacion.objects.create(actividad=actividad, alumno_id=5, calificacion='8.50')

        response = self.client.delete(f'/actividades/{actividad.id}', **self.auth_headers)
        self.assertEqual(response.status_code, 409)
        body = response.json()
        self.assertFalse(body['success'])
        self.assertIn('No se puede eliminar', body['message'])
        self.assertTrue(Actividad.objects.filter(id=actividad.id).exists())


class PonderacionesTests(TestCase):
    def setUp(self):
        self.auth_headers = {'HTTP_AUTHORIZATION': 'Bearer token-valido'}

    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_guardar_ponderaciones_ok(self, mock_validate, mock_materia):
        response = self.client.post(
            '/ponderaciones/10',
            data={
                'ponderaciones': [
                    {'nombre_categoria': 'Examenes', 'porcentaje': '40.00'},
                    {'nombre_categoria': 'Tareas', 'porcentaje': '30.00'},
                    {'nombre_categoria': 'Proyecto', 'porcentaje': '30.00'},
                ]
            },
            content_type='application/json',
            **self.auth_headers,
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['data']['total'], '100.00')
        self.assertEqual(len(body['data']['ponderaciones']), 3)
        mock_validate.assert_called_once_with('token-valido')
        mock_materia.assert_called_once_with(10)

    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_guardar_ponderaciones_rechaza_suma_invalida(self, mock_validate, mock_materia):
        response = self.client.put(
            '/ponderaciones/10',
            data={
                'ponderaciones': [
                    {'nombre_categoria': 'Examenes', 'porcentaje': '40.00'},
                    {'nombre_categoria': 'Tareas', 'porcentaje': '30.00'},
                    {'nombre_categoria': 'Proyecto', 'porcentaje': '20.00'},
                ]
            },
            content_type='application/json',
            **self.auth_headers,
        )

        self.assertEqual(response.status_code, 400)
        body = response.json()
        self.assertFalse(body['success'])
        self.assertIn('100.00', str(body['errors']))

    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=99, rol='docente'))
    def test_guardar_ponderaciones_rechaza_docente_distinto(self, mock_validate, mock_materia):
        response = self.client.post(
            '/ponderaciones/10',
            data={
                'ponderaciones': [
                    {'nombre_categoria': 'Examenes', 'porcentaje': '50.00'},
                    {'nombre_categoria': 'Tareas', 'porcentaje': '50.00'},
                ]
            },
            content_type='application/json',
            **self.auth_headers,
        )

        self.assertEqual(response.status_code, 403)


        class CierreImpresionTests(TestCase):
            def setUp(self):
                self.auth_headers = {'HTTP_AUTHORIZATION': 'Bearer token-valido'}

            @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
            @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
            def test_marcar_imprimir_lista(self, mock_validate, mock_materia):
                response = self.client.post('/materias/10/imprimir-lista', **self.auth_headers)
                self.assertEqual(response.status_code, 200)
                body = response.json()
                self.assertTrue(body['success'])
                estado = EstadoMateria.objects.get(materia_id=10)
                self.assertTrue(estado.lista_impresa)
                mock_validate.assert_called_once_with('token-valido')
                mock_materia.assert_called_once_with(10)
        self.assertFalse(response.json()['success'])

    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_consultar_ponderaciones_ok(self, mock_validate, mock_materia):
        self.client.post(
            '/ponderaciones/10',
            data={
                'ponderaciones': [
                    {'nombre_categoria': 'Examenes', 'porcentaje': '40.00'},
                    {'nombre_categoria': 'Tareas', 'porcentaje': '30.00'},
                    {'nombre_categoria': 'Proyecto', 'porcentaje': '30.00'},
                ]
            },
            content_type='application/json',
            **self.auth_headers,
        )

        response = self.client.get('/ponderaciones/10', **self.auth_headers)
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body['data']['total'], '100.00')
        self.assertEqual(len(body['data']['ponderaciones']), 3)

    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_importar_ponderaciones_excel_ok(self, mock_validate, mock_materia):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['nombre_categoria', 'porcentaje'])
        worksheet.append(['Examenes', 40])
        worksheet.append(['Tareas', 30])
        worksheet.append(['Proyecto', 30])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = self.client.post(
            '/ponderaciones/10/importar',
            data={'archivo': buffer},
            **self.auth_headers,
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['data']['importadas'], 3)


class CalificacionTests(TestCase):
    def setUp(self):
        self.auth_headers = {'HTTP_AUTHORIZATION': 'Bearer token-valido'}
        self.pond = Ponderacion.objects.create(
            materia_id=10,
            nombre_categoria='Exámenes',
            porcentaje='100.00',
        )
        self.actividad = Actividad.objects.create(ponderacion=self.pond, nombre='Examen Final')

    @patch('apps.core.views.is_alumno_en_materia_local', return_value=True)
    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_crear_calificacion_ok(self, mock_validate, mock_materia, mock_is_alumno):
        response = self.client.post(
            '/calificaciones',
            data={
                'actividad_id': self.actividad.id,
                'alumno_id': 5,
                'calificacion': '8.50',
            },
            content_type='application/json',
            **self.auth_headers,
        )

        self.assertEqual(response.status_code, 201)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['data']['calificacion'], '8.50')
        self.assertEqual(body['data']['alumno_id'], 5)
        mock_is_alumno.assert_called_once_with(5, 10)

    @patch('apps.core.views.is_alumno_en_materia_local', return_value=False)
    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_crear_calificacion_alumno_no_inscrito(self, mock_validate, mock_materia, mock_is_alumno):
        response = self.client.post(
            '/calificaciones',
            data={
                'actividad_id': self.actividad.id,
                'alumno_id': 5,
                'calificacion': '8.50',
            },
            content_type='application/json',
            **self.auth_headers,
        )

        self.assertEqual(response.status_code, 400)
        body = response.json()
        self.assertFalse(body['success'])
        self.assertIn('no está inscrito', body['message'])

    @patch('apps.core.views.is_alumno_en_materia_local', return_value=True)
    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_upsert_calificacion(self, mock_validate, mock_materia, mock_is_alumno):
        # Crear primera vez
        response1 = self.client.post(
            '/calificaciones',
            data={
                'actividad_id': self.actividad.id,
                'alumno_id': 5,
                'calificacion': '7.00',
            },
            content_type='application/json',
            **self.auth_headers,
        )
        self.assertEqual(response1.status_code, 201)
        calif_id = response1.json()['data']['id']

        # Actualizar (upsert)
        response2 = self.client.post(
            '/calificaciones',
            data={
                'actividad_id': self.actividad.id,
                'alumno_id': 5,
                'calificacion': '9.00',
            },
            content_type='application/json',
            **self.auth_headers,
        )
        self.assertEqual(response2.status_code, 200)
        body2 = response2.json()
        self.assertEqual(body2['data']['calificacion'], '9.00')

    @patch('apps.core.views.is_alumno_en_materia_local', return_value=True)
    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_editar_calificacion_bloqueado_si_lista_impresa(self, mock_validate, mock_materia, mock_is_alumno):
        # Crear calificación
        from apps.core.models import Calificacion
        calif = Calificacion.objects.create(actividad=self.actividad, alumno_id=5, calificacion='7.00')

        # Marcar lista como impresa
        estado = EstadoMateria.objects.create(materia_id=10, lista_impresa=True)

        # Intentar editar
        response = self.client.put(
            f'/calificaciones/{calif.id}',
            data={'calificacion': '8.00'},
            content_type='application/json',
            **self.auth_headers,
        )

        self.assertEqual(response.status_code, 409)
        body = response.json()
        self.assertFalse(body['success'])
        self.assertIn('lista ya fue impresa', body['message'])

    @patch('apps.core.views.is_alumno_en_materia_local', return_value=True)
    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_importar_calificaciones_ok(self, mock_validate, mock_materia, mock_is_alumno):
        # Crear una actividad
        from openpyxl import Workbook

        workbook = Workbook()
        ws = workbook.active
        ws.append(['matricula', 'actividad_id', 'calificacion'])
        ws.append([5, self.actividad.id, 8.5])
        ws.append([6, self.actividad.id, 7.0])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = self.client.post(f'/calificaciones/importar/10', data={'archivo': buffer}, **self.auth_headers)
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['data']['importadas'], 2)

    @patch('apps.core.views.is_alumno_en_materia_local', return_value=True)
    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_importar_calificaciones_acepta_encabezados_con_acentos(self, mock_validate, mock_materia, mock_is_alumno):
        workbook = Workbook()
        ws = workbook.active
        ws.append(['Matrícula', 'Actividad', 'Calificación'])
        ws.append([5, self.actividad.id, 8.5])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = self.client.post(f'/calificaciones/importar/10', data={'archivo': buffer}, **self.auth_headers)
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['data']['importadas'], 1)

    @patch('apps.core.views.is_alumno_en_materia_local', return_value=True)
    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_importar_calificaciones_bloqueada_si_lista_impresa(self, mock_validate, mock_materia, mock_is_alumno):
        EstadoMateria.objects.create(materia_id=10, lista_impresa=True)

        workbook = Workbook()
        ws = workbook.active
        ws.append(['matricula', 'actividad_id', 'calificacion'])
        ws.append([5, self.actividad.id, 8.5])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = self.client.post(f'/calificaciones/importar/10', data={'archivo': buffer}, **self.auth_headers)
        self.assertEqual(response.status_code, 409)
        body = response.json()
        self.assertFalse(body['success'])
        self.assertIn('lista ya fue impresa', body['message'])

    @patch('apps.core.views.is_alumno_en_materia_local')
    @patch('apps.core.views.get_materia_local', return_value=SimpleNamespace(docente_id=7))
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_importar_calificaciones_con_errores(self, mock_validate, mock_materia, mock_is_alumno):
        # Simular alumno no inscrito en la segunda fila
        mock_is_alumno.side_effect = [True, False]
        workbook = Workbook()
        ws = workbook.active
        ws.append(['matricula', 'actividad_id', 'calificacion'])
        ws.append([5, self.actividad.id, 8.5])
        ws.append([6, self.actividad.id, 12])  # nota inválida

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = self.client.post(f'/calificaciones/importar/10', data={'archivo': buffer}, **self.auth_headers)
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['data']['importadas'], 1)
        self.assertEqual(body['data']['fallos'], 1)


class CalculoPromediosTests(TestCase):
    def setUp(self):
        self.pond_examenes = Ponderacion.objects.create(
            materia_id=10,
            nombre_categoria='Exámenes',
            porcentaje='50.00',
        )
        self.pond_tareas = Ponderacion.objects.create(
            materia_id=10,
            nombre_categoria='Tareas',
            porcentaje='50.00',
        )
        self.act_examen_1 = Actividad.objects.create(ponderacion=self.pond_examenes, nombre='Examen 1')
        self.act_examen_2 = Actividad.objects.create(ponderacion=self.pond_examenes, nombre='Examen 2')
        self.act_tarea_1 = Actividad.objects.create(ponderacion=self.pond_tareas, nombre='Tarea 1')

    def test_redondeo_institucional_bordes(self):
        self.assertEqual(redondear_institucional(Decimal('7.5')), 8)
        self.assertEqual(redondear_institucional(Decimal('7.4')), 7)
        self.assertEqual(redondear_institucional(Decimal('6.0')), 6)
        self.assertEqual(redondear_institucional(Decimal('10.0')), 10)
        self.assertEqual(redondear_institucional(Decimal('7.499999')), 7)
        self.assertEqual(redondear_institucional(Decimal('7.500000')), 8)

    def test_calcular_promedio_ponderado(self):
        Calificacion.objects.create(actividad=self.act_examen_1, alumno_id=5, calificacion='8.00')
        Calificacion.objects.create(actividad=self.act_examen_2, alumno_id=5, calificacion='6.00')
        Calificacion.objects.create(actividad=self.act_tarea_1, alumno_id=5, calificacion='10.00')

        promedio = calcular_promedio_ponderado(5, 10)
        self.assertEqual(promedio.quantize(Decimal('0.01')), Decimal('8.50'))
        self.assertEqual(redondear_institucional(promedio), 9)

    def test_obtener_estadisticas_materia(self):
        Calificacion.objects.create(actividad=self.act_examen_1, alumno_id=5, calificacion='8.00')
        Calificacion.objects.create(actividad=self.act_examen_2, alumno_id=5, calificacion='6.00')
        Calificacion.objects.create(actividad=self.act_tarea_1, alumno_id=5, calificacion='10.00')
        Calificacion.objects.create(actividad=self.act_examen_1, alumno_id=6, calificacion='7.00')
        Calificacion.objects.create(actividad=self.act_examen_2, alumno_id=6, calificacion='7.00')
        Calificacion.objects.create(actividad=self.act_tarea_1, alumno_id=6, calificacion='7.00')

        estadisticas = obtener_estadisticas_materia(10)
        self.assertIsNotNone(estadisticas)
        self.assertEqual(estadisticas['total_alumnos'], 2)
        self.assertEqual(estadisticas['aprobados'], 2)
        self.assertEqual(estadisticas['reprobados'], 0)
        self.assertEqual(estadisticas['promedio_grupal'].quantize(Decimal('0.01')), Decimal('7.75'))


class ConcentradoRestTests(TestCase):
    def setUp(self):
        self.auth_headers = {'HTTP_AUTHORIZATION': 'Bearer token-valido'}
        MateriaProjection.objects.create(
            materia_id=10,
            periodo_id=2,
            nrc='12345',
            nombre='Servicios Web',
            docente_id=7,
            docente_nombre='Docente',
            periodo_nombre='2026-1',
        )
        AlumnoMateriaProjection.objects.create(
            alumno_id=5,
            materia_id=10,
            matricula='20240001',
            nombre='Ana Lopez',
            email='ana@test.local',
            activa=True,
        )
        AlumnoMateriaProjection.objects.create(
            alumno_id=6,
            materia_id=10,
            matricula='20240002',
            nombre='Luis Perez',
            email='luis@test.local',
            activa=True,
        )
        self.pond_examenes = Ponderacion.objects.create(
            materia_id=10,
            nombre_categoria='Exámenes',
            porcentaje='50.00',
        )
        self.pond_tareas = Ponderacion.objects.create(
            materia_id=10,
            nombre_categoria='Tareas',
            porcentaje='50.00',
        )
        self.act_examen_1 = Actividad.objects.create(ponderacion=self.pond_examenes, nombre='Examen 1')
        self.act_examen_2 = Actividad.objects.create(ponderacion=self.pond_examenes, nombre='Examen 2')
        self.act_tarea_1 = Actividad.objects.create(ponderacion=self.pond_tareas, nombre='Tarea 1')

        Calificacion.objects.create(actividad=self.act_examen_1, alumno_id=5, calificacion='8.00')
        Calificacion.objects.create(actividad=self.act_examen_2, alumno_id=5, calificacion='6.00')
        Calificacion.objects.create(actividad=self.act_tarea_1, alumno_id=5, calificacion='10.00')

    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_concentrado_rest_enriquece_con_alumnos_ms3(self, mock_validate):
        response = self.client.get('/concentrado/10', **self.auth_headers)
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['data']['materia_id'], 10)
        self.assertEqual(len(body['data']['categorias']), 2)
        self.assertEqual(len(body['data']['alumnos']), 2)

        alumno_5 = next(item for item in body['data']['alumnos'] if item['alumno_id'] == 5)
        self.assertEqual(alumno_5['matricula'], '20240001')
        self.assertEqual(alumno_5['nombre'], 'Ana Lopez')
        self.assertEqual(alumno_5['promedio_real'], '8.50')
        self.assertEqual(alumno_5['promedio_redondeado'], 9)

        alumno_6 = next(item for item in body['data']['alumnos'] if item['alumno_id'] == 6)
        self.assertEqual(alumno_6['matricula'], '20240002')
        self.assertEqual(alumno_6['nombre'], 'Luis Perez')
        self.assertEqual(alumno_6['promedio_real'], '0.00')
        self.assertEqual(alumno_6['promedio_redondeado'], 0)


class CerrarMateriaTests(TestCase):
    def setUp(self):
        MateriaProjection.objects.create(
            materia_id=10,
            periodo_id=2,
            nrc='12345',
            nombre='Servicios Web',
            docente_id=7,
            docente_nombre='Docente',
            periodo_nombre='2026-1',
        )
        MateriaProjection.objects.create(
            materia_id=11,
            periodo_id=2,
            nrc='54321',
            nombre='Redes',
            docente_id=7,
            docente_nombre='Docente',
            periodo_nombre='2026-1',
        )

    @override_settings(USE_EVENT_BUS=True)
    @patch('apps.core.views.publish_materia_calificaciones_cerradas')
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_cerrar_materia_ok(self, _mock_auth, mock_publish):
        response = self.client.post(
            '/materias/10/cerrar',
            HTTP_AUTHORIZATION='Bearer token-valido',
        )
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertTrue(body['data']['evento_publicado'])
        mock_publish.assert_called_once()
        estado = EstadoMateria.objects.get(materia_id=10)
        self.assertTrue(estado.cerrada)
        self.assertTrue(estado.notificacion_enviada)

    @override_settings(USE_EVENT_BUS=False)
    @patch('apps.core.views.publish_materia_calificaciones_cerradas')
    @patch('apps.core.views.validate_access_token', return_value=SimpleNamespace(user_id=7, rol='docente'))
    def test_cerrar_materia_sin_event_bus(self, _mock_auth, mock_publish):
        response = self.client.post(
            '/materias/11/cerrar',
            HTTP_AUTHORIZATION='Bearer token-valido',
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(EstadoMateria.objects.get(materia_id=11).cerrada)
        self.assertFalse(response.json()['data']['evento_publicado'])
        mock_publish.assert_called_once()
